import smtplib
import os
from traceback import format_exc
from datetime import date
from datetime import datetime
from datetime import timedelta
import time
import datetime
import openpyxl.utils.datetime
import pandas.io.sql
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
import pyodbc
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from smtplib import SMTP
import pandas as pd
import xlsxwriter
import logging
import tracestack as ts

enrollments = """ query1 """

appointments = """ query2 """

directory_name = r'directory_name'

through_time = date.today()
start_time = date.today() - timedelta(days=1)

excel_file_path = 'excel_builder_template' + str(start_time) + '.xlsx'
email_file = directory_name + '/' + excel_file_path

email_list = ['trashpanda@email.com', i, i, i, 'etc.']

# creates log file
logging.basicConfig(
    filename=r'directory_name' + '/' + str(through_time) + ' log' + '.log',
    level=logging.INFO,
    format='%(asctime)s:%(levelname)s:%(message)s')

log_file = r'directory_name' + '/' + str(through_time) + ' log' + '.log'

# establishes connection to SQL Server
cursor = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};'
                        'SERVER=Server_Name;'
                        'DATABASE="Database_Name";'
                        'UID="User_ID";'
                        'PWD="Database_Password"')


# funtion that drops duplicate entries with defined subsets.
def drop_duplicates(df1, df2):
    df1.drop_duplicates(subset=['enrollee_name', 'program_name'], keep='first', inplace=True)
    df2.drop_duplicates(subset=['client_full_name', 'appointment_date', 'event_name'], keep='first', inplace=True)


# function that builds directory where files are saved (dir_name
def create_directory(directory_name):
     if not os.path.exists(directory_name):
         os.makedirs(directory_name)
     return directory_name


# email function that builds email. (files, recipients, subject, body)
def send_email(file, recipients, subject, body):
    emailfrom = "email_account@gmail.com"
    subject = 'Subject of Email'
    body = 'Body of Email |' \
           '***This report was generated automatically*** \
\n\n\nThe attached report includes all information pertaining to query parameters. All data is exported from database.\
  If there are any issues \
or if anything appears to be an error please contact "emailfrom".' \
           '' \
           '\n\nThank you!'

    msg = MIMEMultipart()
    msg["From"] = emailfrom
    msg["To"] = ",".join(recipients)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, 'plain'))

    filename = excel_file_path
    attachment = open(str(filename), 'rb')

    part = MIMEBase('application', "octet-stream")
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment; filename= " + filename)

    msg.attach(part)
    text = msg.as_string()
    server = SMTP("SMTP Server_Name")
    server.sendmail(emailfrom, recipients, msg.as_string())
    server.quit()


# email function that attaches the log file and error message.
def error_email(files, recipients, subject, body):
    emailfrom = "email_account@gmail"
    subject = 'ERROR: Error message'
    body = body

    msg = MIMEMultipart()
    msg["From"] = emailfrom
    msg["To"] = ",".join(recipients)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, 'plain'))

    filename = log_file

    attachment = open(str(filename), 'rb')

    part = MIMEBase('application', "octet-stream")
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment; filename= " + filename)

    msg.attach(part)
    text = msg.as_string()
    server = SMTP("SMTP Server Name")
    server.sendmail(emailfrom, recipients, msg.as_string())
    server.quit()


# formats the sheet using XlsxWriter and Pandas
def format_sheet():
    writer = pd.ExcelWriter(excel_file_path, engine='xlsxwriter')

    # build dataframes from SQL Queries
    df1 = pd.read_sql(query1, cursor)
    df2 = pd.read_sql(query2, cursor)

    # calls the drop duplicates function with the two dataframes as parameters.
    drop_duplicates(df1, df2)

    # exports formatted dataframe to Excel
    df1.to_excel(writer, index=False, sheet_name='Sheet1_Name')
    df2.to_excel(writer, index=False, sheet_name='Sheet2_Name')

    # saves file to excel_file_path document.
    writer.save()


def main():
    # calls the format function
    format_sheet()

    # loads excel_file_path to save to a different directory. Then closes the file
    directory_file = load_workbook(excel_file_path)
    directory_file.save(directory_name + '/' + excel_file_path)
    directory_file.close()

    # calls send_email function
    send_email(email_file, email_list, '', '')


def execute():
    try:
        main()
        logging.info('Result: Success. Program has run correctly.')
    except pyodbc.ProgrammingError:
        logging.info('Result: Failure. Program has encountered errors. Please troubleshoot program for errors.' + str(
            format_exc()))
        error_email(filename, email_list, '', str(format_exc()))
    except pandas.io.sql.DatabaseError:
        logging.info('Result: Failure. Program has encountered errors. Please troubleshoot program for errors. ' + str(
            format_exc()))
        error_email(filename, email_list, '', str(format_exc()))
    except OSError:
        logging.info('Result: Failure. Program has encountered errors. Please troubleshoot program for errors. ' + str(
            format_exc()))
        error_email(filename, email_list, '', str(format_exc()))
    else:
        logging.info(
            'Result: Failure. Program has encountered errors other than the exceptions. Please troubleshoot program for unidentified errors.' + str(
                format_exc()))
        error_email(filename, email_list, '', str(format_exc()))
    finally:
        # closes connection to SQL Server
        cursor.close()

# calls the execute function        
execute()
